from __future__ import annotations

import argparse
import json
import re
import shutil
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


RECTANGULAR_CELL_LIMIT = 250_000
SPARSE_CELL_LIMIT = 500_000


def display(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).replace("\r\n", "\n").replace("\r", "\n")
    return text.replace("|", "\\|").replace("\n", "<br>")


def instantiated_cells(ws) -> list[Any]:
    cells = getattr(ws, "_cells", None)
    if isinstance(cells, dict):
        return list(cells.values())
    return [cell for row in ws.iter_rows() for cell in row]


def populated_cells(ws) -> list[Any]:
    return [cell for cell in instantiated_cells(ws) if cell.value is not None]


def safe_sheet_folder(index: int, title: str) -> str:
    safe = re.sub(r'[<>:"/\\|?*\x00-\x1f]', "_", title).strip().rstrip(".")
    safe = safe[:60] or "sheet"
    return f"{index:02d}-{safe}"


def rectangular_table(ws, max_row: int, max_col: int) -> str:
    rows = ["| Row | " + " | ".join(get_column_letter(i) for i in range(1, max_col + 1)) + " |"]
    rows.append("| ---: | " + " | ".join("---" for _ in range(max_col)) + " |")
    for row_idx in range(1, max_row + 1):
        values = [display(ws.cell(row_idx, col_idx).value) for col_idx in range(1, max_col + 1)]
        rows.append(f"| {row_idx} | " + " | ".join(values) + " |")
    return "\n".join(rows) + "\n"


def sparse_table(cells: Iterable[Any]) -> str:
    rows = ["| Cell | Value |", "| --- | --- |"]
    for cell in sorted(cells, key=lambda item: (item.row, item.column)):
        rows.append(f"| {cell.coordinate} | {display(cell.value)} |")
    return "\n".join(rows) + "\n"


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("input_workbook")
    parser.add_argument("output_dir")
    args = parser.parse_args()

    source = Path(args.input_workbook).resolve()
    output = Path(args.output_dir).resolve()
    output.mkdir(parents=True, exist_ok=True)
    shutil.copy2(source, output / source.name)

    keep_vba = source.suffix.lower() in {".xlsm", ".xltm"}
    formulas_book = load_workbook(source, data_only=False, read_only=False, keep_vba=keep_vba, keep_links=True)
    values_book = load_workbook(source, data_only=True, read_only=False, keep_vba=keep_vba, keep_links=True)
    sheets_root = output / "sheets-data"
    sheets_root.mkdir(parents=True, exist_ok=True)

    values_index = [f"# Displayed values index — {source.name}", ""]
    formulas_index = [f"# Formulas index — {source.name}", ""]
    info_parts = [f"# Workbook information — {source.name}", ""]
    total_formulas = 0
    total_errors = 0
    quality_warnings: list[str] = []
    sheet_metrics: list[dict[str, Any]] = []

    external_links = len(getattr(formulas_book, "_external_links", []))
    if external_links:
        quality_warnings.append(f"Workbook contains {external_links} external link(s).")

    for index, formula_ws in enumerate(formulas_book.worksheets, 1):
        value_ws = values_book[formula_ws.title]
        formula_cells_all = populated_cells(formula_ws)
        value_cells_all = populated_cells(value_ws)
        max_row = max((cell.row for cell in formula_cells_all), default=0)
        max_col = max((cell.column for cell in formula_cells_all), default=0)
        min_row = min((cell.row for cell in formula_cells_all), default=0)
        min_col = min((cell.column for cell in formula_cells_all), default=0)
        populated_row_span = max_row - min_row + 1 if max_row else 0
        populated_col_span = max_col - min_col + 1 if max_col else 0
        rectangular_size = max_row * max_col
        folder_name = safe_sheet_folder(index, formula_ws.title)
        sheet_dir = sheets_root / folder_name
        sheet_dir.mkdir(parents=True, exist_ok=True)

        if len(formula_cells_all) > SPARSE_CELL_LIMIT:
            raise RuntimeError(
                f"Sheet {formula_ws.title!r} contains more than {SPARSE_CELL_LIMIT:,} populated/stored cells; "
                "refine the workbook before packaging."
            )

        use_sparse = rectangular_size > RECTANGULAR_CELL_LIMIT
        if use_sparse:
            quality_warnings.append(
                f"Sheet {formula_ws.title!r} uses sparse output because its rectangular range contains {rectangular_size:,} cells."
            )
            values_body = sparse_table(value_cells_all)
        elif max_row and max_col:
            values_body = rectangular_table(value_ws, max_row, max_col)
        else:
            values_body = "_Empty sheet._\n"

        formulas = []
        cached_errors = []
        for cell in formula_cells_all:
            if cell.data_type == "f" or (isinstance(cell.value, str) and cell.value.startswith("=")):
                cached = value_ws[cell.coordinate].value
                formulas.append((cell.coordinate, cell.value, cached, cell.number_format))
                if isinstance(cached, str) and cached.startswith("#"):
                    cached_errors.append((cell.coordinate, cached))

        total_formulas += len(formulas)
        total_errors += len(cached_errors)
        hidden_rows = sum(1 for dim in formula_ws.row_dimensions.values() if dim.hidden)
        hidden_cols = sum(1 for dim in formula_ws.column_dimensions.values() if dim.hidden)
        charts = len(getattr(formula_ws, "_charts", []))
        images = len(getattr(formula_ws, "_images", []))

        values_path = sheet_dir / "values.md"
        formulas_path = sheet_dir / "formulas.md"
        values_path.write_text(f"# Values — {formula_ws.title}\n\n{values_body}", encoding="utf-8")
        formula_lines = [f"# Formulas — {formula_ws.title}", ""]
        if formulas:
            formula_lines.extend(["| Cell | Formula | Cached result | Number format |", "| --- | --- | --- | --- |"])
            formula_lines.extend(
                f"| {coordinate} | {display(formula)} | {display(cached)} | {display(number_format)} |"
                for coordinate, formula, cached, number_format in formulas
            )
        else:
            formula_lines.append("_No formulas._")
        formulas_path.write_text("\n".join(formula_lines) + "\n", encoding="utf-8")

        relative_values = f"sheets-data/{folder_name}/values.md"
        relative_formulas = f"sheets-data/{folder_name}/formulas.md"
        values_index.append(f"- [{index}. {formula_ws.title}]({relative_values})")
        formulas_index.append(f"- [{index}. {formula_ws.title}]({relative_formulas}) — {len(formulas)} formula(s)")
        info_parts.extend(
            [
                f"## {index}. {formula_ws.title}",
                f"- Visibility: {formula_ws.sheet_state}",
                f"- Populated cells: {len(formula_cells_all)}",
                f"- Populated bounds: A1:{get_column_letter(max_col)}{max_row}" if max_row and max_col else "- Populated bounds: empty",
                f"- Output mode: {'sparse' if use_sparse else 'rectangular'}",
                f"- Formulas: {len(formulas)}",
                f"- Cached formula errors: {len(cached_errors)}",
                f"- Merged ranges: {len(formula_ws.merged_cells.ranges)}",
                f"- Hidden rows / columns: {hidden_rows} / {hidden_cols}",
                f"- Charts / embedded images: {charts} / {images}",
                f"- Values: [{relative_values}]({relative_values})",
                f"- Formulas: [{relative_formulas}]({relative_formulas})",
                "",
            ]
        )
        sheet_metrics.append(
            {
                "index": index,
                "title": formula_ws.title,
                "visibility": formula_ws.sheet_state,
                "populated_cells": len(formula_cells_all),
                "max_row": max_row,
                "max_column": max_col,
                "min_row": min_row,
                "min_column": min_col,
                "populated_row_span": populated_row_span,
                "populated_column_span": populated_col_span,
                "output_mode": "sparse" if use_sparse else "rectangular",
                "formulas": len(formulas),
                "cached_formula_errors": len(cached_errors),
                "hidden_rows": hidden_rows,
                "hidden_columns": hidden_cols,
                "charts": charts,
                "images": images,
                "merged_ranges": len(formula_ws.merged_cells.ranges),
            }
        )

    summary = [
        f"- Sheets: {len(formulas_book.worksheets)}",
        f"- Defined names: {len(formulas_book.defined_names)}",
        f"- External links: {external_links}",
        f"- Calculation mode: {formulas_book.calculation.calcMode or 'unspecified'}",
        f"- Total formulas: {total_formulas}",
        f"- Cached formula errors: {total_errors}",
        "",
    ]
    info_parts[2:2] = summary
    (output / "values.md").write_text("\n".join(values_index) + "\n", encoding="utf-8")
    (output / "formulas.md").write_text("\n".join(formulas_index) + "\n", encoding="utf-8")
    (output / "workbook-info.md").write_text("\n".join(info_parts), encoding="utf-8")

    quality = [
        f"# Quality report — {source.name}",
        "",
        f"- Sheets analyzed: {len(sheet_metrics)}",
        f"- Total formulas: {total_formulas}",
        f"- Cached formula errors: {total_errors}",
        f"- External links: {external_links}",
        "",
        "## Warnings",
        "",
    ]
    quality.extend(f"- {warning}" for warning in quality_warnings)
    if not quality_warnings:
        quality.append("- No structural warnings detected.")
    quality.extend(["", "Cached values may be stale if Excel did not recalculate and save the workbook before packaging.", ""])
    (output / "quality-report.md").write_text("\n".join(quality), encoding="utf-8")
    (output / "excel-metrics.json").write_text(
        json.dumps(
            {
                "sheet_count": len(sheet_metrics),
                "total_formulas": total_formulas,
                "cached_formula_errors": total_errors,
                "external_links": external_links,
                "warnings": quality_warnings,
                "sheets": sheet_metrics,
            },
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )
    formulas_book.close()
    values_book.close()
    print(f"Extracted {len(sheet_metrics)} sheets and {total_formulas} formulas")


if __name__ == "__main__":
    main()
