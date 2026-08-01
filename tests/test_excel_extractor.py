from __future__ import annotations

import json
import subprocess
import sys
import tempfile
import unittest
import zipfile
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference


ROOT = Path(__file__).resolve().parents[1]
EXTRACTOR = ROOT / "extract-excel-package.py"


class ExcelExtractorTests(unittest.TestCase):
    def run_extractor(self, workbook_path: Path, output_path: Path) -> None:
        subprocess.run([sys.executable, str(EXTRACTOR), str(workbook_path), str(output_path)], check=True)

    def test_splits_values_and_formulas_per_sheet(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            root = Path(temporary)
            source = root / "sample.xlsx"
            output = root / "package"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Summary"
            sheet["A1"] = 4
            sheet["B1"] = 5
            sheet["C1"] = "=A1+B1"
            workbook.save(source)
            self.run_extractor(source, output)

            metrics = json.loads((output / "excel-metrics.json").read_text(encoding="utf-8"))
            self.assertEqual(metrics["sheet_count"], 1)
            self.assertEqual(metrics["total_formulas"], 1)
            self.assertEqual(metrics["sheets"][0]["min_row"], 1)
            self.assertEqual(metrics["sheets"][0]["min_column"], 1)
            self.assertEqual(metrics["sheets"][0]["populated_column_span"], 3)
            self.assertEqual(metrics["sheets"][0]["merged_ranges"], 0)
            self.assertTrue((output / "sheets-data" / "01-Summary" / "values.md").exists())
            self.assertIn("=A1+B1", (output / "sheets-data" / "01-Summary" / "formulas.md").read_text(encoding="utf-8"))

    def test_extreme_dimension_uses_sparse_output(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            root = Path(temporary)
            source = root / "sparse.xlsx"
            output = root / "package"
            workbook = Workbook()
            sheet = workbook.active
            sheet["A1"] = "start"
            sheet["XFD1048576"] = "end"
            workbook.save(source)
            self.run_extractor(source, output)

            metrics = json.loads((output / "excel-metrics.json").read_text(encoding="utf-8"))
            self.assertEqual(metrics["sheets"][0]["output_mode"], "sparse")
            self.assertEqual(metrics["sheets"][0]["populated_column_span"], 16_384)
            values = (output / "sheets-data" / "01-Sheet" / "values.md").read_text(encoding="utf-8")
            self.assertIn("XFD1048576", values)
            self.assertLess(len(values), 10_000)

    def test_reports_drawing_risk_metrics(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            root = Path(temporary)
            source = root / "chart.xlsx"
            output = root / "package"
            workbook = Workbook()
            sheet = workbook.active
            for row in range(1, 5):
                sheet.append([f"Row {row}", row])
            chart = BarChart()
            chart.add_data(Reference(sheet, min_col=2, min_row=1, max_row=4))
            sheet.add_chart(chart, "D2")
            workbook.save(source)
            self.run_extractor(source, output)

            metrics = json.loads((output / "excel-metrics.json").read_text(encoding="utf-8"))
            self.assertEqual(metrics["sheets"][0]["charts"], 1)
            self.assertEqual(metrics["sheets"][0]["images"], 0)

    def test_missing_calculation_properties_are_reported_as_unspecified(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            root = Path(temporary)
            original = root / "original.xlsx"
            source = root / "without-calculation-properties.xlsx"
            output = root / "package"
            workbook = Workbook()
            workbook.active["A1"] = "valid workbook without calcPr"
            workbook.save(original)

            with zipfile.ZipFile(original) as input_archive, zipfile.ZipFile(source, "w") as output_archive:
                for item in input_archive.infolist():
                    content = input_archive.read(item.filename)
                    if item.filename == "xl/workbook.xml":
                        text = content.decode("utf-8")
                        text = text.replace(
                            '<calcPr calcId="124519" fullCalcOnLoad="1"/>',
                            "",
                        )
                        content = text.encode("utf-8")
                    output_archive.writestr(item, content)

            self.run_extractor(source, output)

            info = (output / "workbook-info.md").read_text(encoding="utf-8")
            self.assertIn("- Calculation mode: unspecified", info)


if __name__ == "__main__":
    unittest.main()
